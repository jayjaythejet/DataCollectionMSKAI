"""
Excel read/write handler using openpyxl.
Reads accession numbers from the user-selected input file (never modified).
Writes all answers to responses_<input_name>.xlsx in the same folder as the input file.
"""
import os
import openpyxl
import xlrd
from datetime import datetime

# All columns the app writes (in addition to reading 'accession')
ANSWER_COLUMNS = [
    # Section 1: Image Quality (Likert 1-5)
    "contrast_resolution", "edge_sharpness", "fat_suppression",
    "fluid_brightness", "image_noise", "motion_artifact",
    "partial_volume_effects", "overall_image_quality",
    # Section 2: Structure Visibility (Likert 1-5)
    "acl_visibility", "pcl_visibility", "mcl_visibility", "lcl_visibility",
    "medial_meniscus_visibility", "lateral_meniscus_visibility",
    "extensor_tendons_visibility", "articular_cartilage_visibility",
    "bones_visibility",
    # Section 3: Pathology (Yes/No or No/Partial/Full)
    "acl_tear", "pcl_tear", "mcl_tear", "lcl_tear",
    "medial_meniscus_tear", "lateral_meniscus_tear",
    "articular_cartilage_defect", "bone_marrow_edema",
    # Extras
    "notes", "status",
]


def _get_app_data_dir() -> str:
    """Return the writable app data directory, cross-platform."""
    app_name = "MSK_DataCollector"
    try:
        from platformdirs import user_data_dir
        base = user_data_dir(app_name, appauthor=False)
    except ImportError:
        if os.name == "nt":
            base = os.path.join(
                os.environ.get("LOCALAPPDATA", os.path.expanduser("~")), app_name
            )
        else:
            base = os.path.join(os.path.expanduser("~"), ".local", "share", app_name)
    os.makedirs(base, exist_ok=True)
    return base


class ExcelHandler:
    def __init__(self, answer_columns=None):
        self.answer_columns = answer_columns or ANSWER_COLUMNS
        self.in_filepath = None
        self.in_workbook = None
        self.in_sheet = None
        self.accession_col = None
        self.data_rows = []       # row numbers in input sheet (1-based)
        self.accession_list = []  # ordered list of accession strings (works for both .xls and .xlsx)

        self.out_path = None
        self.out_workbook = None
        self.out_sheet = None
        self.out_col_index = {}   # col_name -> col_number (1-based) in output sheet
        self.out_data_rows = []   # output row numbers indexed by record index

    def load(self, filepath: str) -> list[str]:
        """
        Load the input workbook (read-only intent — never saved back).
        Returns list of accession values in order.
        Raises ValueError if no 'accession' column found.
        """
        self.in_filepath = filepath
        ext = os.path.splitext(filepath)[1].lower()

        if ext == ".xls":
            # Old binary Excel format — use xlrd (read-only)
            wb_xls = xlrd.open_workbook(filepath)
            sh = wb_xls.sheet_by_index(0)
            headers = {}
            for col in range(sh.ncols):
                v = sh.cell_value(0, col)
                if v:
                    headers[str(v).strip().lower()] = col  # 0-based
            if "accession" not in headers:
                raise ValueError(
                    "No 'accession' column found in the Excel file. "
                    "Please ensure the header row has a column named 'accession'."
                )
            acc_col = headers["accession"]
            accessions = []
            for row in range(1, sh.nrows):  # skip header row
                v = sh.cell_value(row, acc_col)
                if v is not None and str(v).strip():
                    accessions.append(str(v).strip())
            # in_workbook/in_sheet remain None — not needed after load for .xls
        else:
            # .xlsx / .xlsm — use openpyxl
            try:
                self.in_workbook = openpyxl.load_workbook(filepath)
            except Exception as e:
                msg = str(e)
                if "not a zip file" in msg.lower() or "bad magic" in msg.lower():
                    raise ValueError(
                        "Could not read this file. It may be in an unsupported format "
                        "or password-protected. Try saving it as .xlsx and re-opening."
                    ) from e
                raise
            self.in_sheet = self.in_workbook.active
            headers = {}
            for cell in self.in_sheet[1]:
                if cell.value is not None:
                    headers[str(cell.value).strip().lower()] = cell.column
            if "accession" not in headers:
                raise ValueError(
                    "No 'accession' column found in the Excel file. "
                    "Please ensure the header row has a column named 'accession'."
                )
            self.accession_col = headers["accession"]
            accessions = []
            for row in range(2, self.in_sheet.max_row + 1):
                val = self.in_sheet.cell(row=row, column=self.accession_col).value
                if val is not None and str(val).strip():
                    accessions.append(str(val).strip())

        if not accessions:
            raise ValueError("The Excel file has no accession data rows.")

        self.accession_list = accessions
        input_dir = os.path.dirname(os.path.abspath(filepath))
        input_stem = os.path.splitext(os.path.basename(filepath))[0]
        self._ensure_output_workbook(accessions, output_dir=input_dir, output_stem=input_stem)
        return accessions

    def _ensure_output_workbook(self, accessions: list, output_dir: str = None, output_stem: str = "responses"):
        """Create or load the output workbook next to the input file."""
        if output_dir is None:
            output_dir = _get_app_data_dir()
        self.out_path = os.path.join(output_dir, f"responses_{output_stem}.xlsx")
        out_headers = ["accession"] + self.answer_columns

        if os.path.exists(self.out_path):
            self.out_workbook = openpyxl.load_workbook(self.out_path)
            self.out_sheet = self.out_workbook.active

            # Read existing headers
            existing_headers = {}
            for cell in self.out_sheet[1]:
                if cell.value is not None:
                    existing_headers[str(cell.value).strip().lower()] = cell.column

            # Add any missing columns
            next_col = self.out_sheet.max_column + 1
            for col_name in out_headers:
                if col_name not in existing_headers:
                    self.out_sheet.cell(row=1, column=next_col, value=col_name)
                    existing_headers[col_name] = next_col
                    next_col += 1

            self.out_col_index = existing_headers

            # Build accession -> row map for existing output rows
            acc_col = self.out_col_index["accession"]
            existing_rows = {}
            for r in range(2, self.out_sheet.max_row + 1):
                v = self.out_sheet.cell(row=r, column=acc_col).value
                if v is not None:
                    existing_rows[str(v).strip()] = r

            # Ensure each input accession has a row in output
            next_out_row = max(existing_rows.values(), default=1) + 1
            if next_out_row < 2:
                next_out_row = 2

            self.out_data_rows = []
            for acc in accessions:
                if acc in existing_rows:
                    self.out_data_rows.append(existing_rows[acc])
                else:
                    self.out_sheet.cell(row=next_out_row, column=acc_col, value=acc)
                    existing_rows[acc] = next_out_row
                    self.out_data_rows.append(next_out_row)
                    next_out_row += 1
        else:
            # Create fresh output workbook
            self.out_workbook = openpyxl.Workbook()
            self.out_sheet = self.out_workbook.active
            for col, name in enumerate(out_headers, 1):
                self.out_sheet.cell(row=1, column=col, value=name)
            self.out_col_index = {name: col for col, name in enumerate(out_headers, 1)}

            acc_col = self.out_col_index["accession"]
            self.out_data_rows = []
            for i, acc in enumerate(accessions):
                out_row = i + 2
                self.out_sheet.cell(row=out_row, column=acc_col, value=acc)
                self.out_data_rows.append(out_row)

        self.out_workbook.save(self.out_path)

    def read_row(self, row_index: int) -> dict:
        """
        Read all answer column values for a given row_index (0-based).
        Reads from the output workbook.
        """
        row = self.out_data_rows[row_index]
        result = {}
        for col_name in self.answer_columns:
            col_num = self.out_col_index.get(col_name)
            if col_num:
                result[col_name] = self.out_sheet.cell(row=row, column=col_num).value
            else:
                result[col_name] = None
        return result

    def write_row(self, row_index: int, answers: dict):
        """
        Write answers to the output workbook only. Never touches the input file.
        """
        row = self.out_data_rows[row_index]
        for col_name, value in answers.items():
            col_num = self.out_col_index.get(col_name)
            if col_num:
                self.out_sheet.cell(row=row, column=col_num, value=value)
        self.out_workbook.save(self.out_path)

    def get_completion_status(self) -> dict:
        """Returns dict of {row_index: status_string} for all rows."""
        status_col = self.out_col_index.get("status")
        result = {}
        for i, row in enumerate(self.out_data_rows):
            if status_col:
                val = self.out_sheet.cell(row=row, column=status_col).value
                result[i] = val or "incomplete"
            else:
                result[i] = "incomplete"
        return result

    def get_accession(self, row_index: int) -> str:
        return self.accession_list[row_index]

    def get_output_path(self) -> str:
        """Return the path where responses are being saved."""
        return self.out_path or ""

    def total_records(self) -> int:
        return len(self.accession_list)
